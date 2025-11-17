# academics/views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.db import transaction
from django.db.models import Q, Avg, Max, Min, Count
from django.core.paginator import Paginator
import json
from accounts.models import User
from .models import Subject, ClassLevel, AcademicYear, Term, ClassSubject, Result
import ast
from django.utils import timezone
from django.db import models



@login_required
def subject_list(request):
    """List all subjects with filtering and pagination"""
    subjects = Subject.objects.all()
    
    category = request.GET.get('category')
    if category:
        subjects = subjects.filter(category=category)
    
    is_active = request.GET.get('is_active')
    if is_active:
        subjects = subjects.filter(is_active=is_active == 'true')
    
    search = request.GET.get('search')
    if search:
        subjects = subjects.filter(
            Q(name__icontains=search) |
            Q(code__icontains=search) |
            Q(description__icontains=search)
        )
    
    total_subjects = subjects.count()
    core_subjects = subjects.filter(category='core').count()
    elective_subjects = subjects.filter(category='elective').count()
    active_subjects = subjects.filter(is_active=True).count()
    
    paginator = Paginator(subjects, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'category': category,
        'is_active': is_active,
        'search': search,
        'total_subjects': total_subjects,
        'core_subjects': core_subjects,
        'elective_subjects': elective_subjects,
        'active_subjects': active_subjects,
        'categories': Subject.SUBJECT_CATEGORY_CHOICES,
    }

    return render(request, 'pages/admin_dashboard/subject_lists.html', {'context': context})



@login_required
@require_http_methods(["POST"])
@transaction.atomic
def subject_create(request):
    """API endpoint for creating subjects"""
    try:
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            data = request.POST
        
        name = data.get("name")
        code = data.get("code")
        description = data.get("description")
        category = data.get("category", "core")
        is_active = data.get("is_active") == "true" or data.get("is_active") == "on"
        teacher_id = data.get("teacher")
        
        if not all([name, code]):
            return JsonResponse({
                'success': False, 
                'error': 'Name and code are required.'
            }, status=400)
        
        if Subject.objects.filter(code=code).exists():
            return JsonResponse({
                'success': False, 
                'error': 'A subject with this code already exists.'
            }, status=400)
        
        subject = Subject.objects.create(
            name=name,
            code=code,
            description=description,
            category=category,
            is_active=is_active,
            teacher_id=teacher_id if teacher_id else None,
        )
        
        response_data = {
            'success': True,
            'message': f'Subject {name} created successfully.'
        }
        
        if not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            messages.success(request, response_data['message'])
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'An error occurred while creating subject: {str(e)}'
        }, status=400)


@login_required
@require_http_methods(["POST"])
@transaction.atomic
def subject_update(request, subject_id):
    """API endpoint for updating subjects"""
    try:
        subject = get_object_or_404(Subject, id=subject_id)
        
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            data = request.POST
        
        name = data.get("name")
        code = data.get("code")
        description = data.get("description")
        category = data.get("category")
        is_active = data.get("is_active") == "true" or data.get("is_active") == "on"
        teacher_id = data.get("teacher")
        
        if code != subject.code and Subject.objects.filter(code=code).exists():
            return JsonResponse({
                'success': False, 
                'error': 'A subject with this code already exists.'
            }, status=400)
        
        subject.name = name
        subject.code = code
        subject.description = description
        subject.category = category
        subject.is_active = is_active
        subject.teacher_id = teacher_id if teacher_id else None
        subject.save()
        
        response_data = {
            'success': True,
            'message': f'Subject {name} updated successfully.',
        }
        
        if not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            messages.success(request, response_data['message'])
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'An error occurred while updating subject: {str(e)}'
        }, status=400)


@login_required
@require_http_methods(["DELETE", "POST"])
def subject_delete(request, subject_id):
    """Delete a subject"""
    try:
        subject = get_object_or_404(Subject, id=subject_id)
        subject_name = subject.name
        subject.delete()
        
        response_data = {
            'success': True,
            'message': f'Subject {subject_name} deleted successfully.',
        }
        
        if not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            messages.success(request, response_data['message'])
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'An error occurred while deleting subject: {str(e)}'
        }, status=400)



@login_required
def class_list(request):
    """List all classes with filtering and pagination"""
    classes = ClassLevel.objects.all()
    
    is_active = request.GET.get('is_active')
    if is_active:
        classes = classes.filter(is_active=is_active == 'true')
    
    search = request.GET.get('search')
    if search:
        classes = classes.filter(
            Q(name__icontains=search) |
            Q(code__icontains=search) |
            Q(description__icontains=search)
        )
    
    total_classes = classes.count()
    active_classes = classes.filter(is_active=True).count()
    total_capacity = sum(cls.capacity for cls in classes)
    total_students = sum(cls.current_students_count for cls in classes)
    
    paginator = Paginator(classes, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'is_active': is_active,
        'search': search,
        'total_classes': total_classes,
        'active_classes': active_classes,
        'total_capacity': total_capacity,
        'total_students': total_students,
    }
    return render(request, 'pages/admin_dashboard/class_list.html', {'context': context})


@login_required
@require_http_methods(["POST"])
@transaction.atomic
def class_create(request):
    """API endpoint for creating classes"""
    try:
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            data = request.POST
        
        name = data.get("name")
        code = data.get("code")
        description = data.get("description")
        capacity = data.get("capacity", 30)
        display_order = data.get("display_order", 0)
        is_active = data.get("is_active") == "true" or data.get("is_active") == "on"
        form_teacher_id = data.get("form_teacher")
        subject_ids = data.getlist("subjects") if hasattr(data, 'getlist') else data.get("subjects", [])
        
        if not name:
            return JsonResponse({
                'success': False, 
                'error': 'Name is required.'
            }, status=400)
        
        if ClassLevel.objects.filter(name=name).exists():
            return JsonResponse({
                'success': False, 
                'error': 'A class with this name already exists.'
            }, status=400)
        
        if code and ClassLevel.objects.filter(code=code).exists():
            return JsonResponse({
                'success': False, 
                'error': 'A class with this code already exists.'
            }, status=400)
        
        class_level = ClassLevel.objects.create(
            name=name,
            code=code,
            description=description,
            capacity=capacity,
            display_order=display_order,
            is_active=is_active,
            form_teacher_id=form_teacher_id if form_teacher_id else None,
        )
        
        if subject_ids:
            if isinstance(subject_ids, str):
                import ast
                subject_ids = ast.literal_eval(subject_ids)
            subjects = Subject.objects.filter(id__in=subject_ids)
            class_level.subjects.set(subjects)
        
        response_data = {
            'success': True,
            'message': f'Class {name} created successfully.',
            'class_level': {
                'id': str(class_level.id),
                'name': class_level.name,
                'code': class_level.code,
                'capacity': class_level.capacity,
                'is_active': class_level.is_active,
            }
        }
        
        if not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            messages.success(request, response_data['message'])
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'An error occurred while creating class: {str(e)}'
        }, status=400)



@login_required
@require_http_methods(["POST"])
@transaction.atomic
def class_update(request, class_id):
    """API endpoint for updating classes"""
    try:
        class_level = get_object_or_404(ClassLevel, id=class_id)
        
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            data = request.POST
        
        name = data.get("name")
        code = data.get("code")
        description = data.get("description")
        capacity = data.get("capacity")
        display_order = data.get("display_order")
        is_active = data.get("is_active") == "true" or data.get("is_active") == "on"
        form_teacher_id = data.get("form_teacher")
        subject_ids = data.getlist("subjects") if hasattr(data, 'getlist') else data.get("subjects", [])
        
        if name != class_level.name and ClassLevel.objects.filter(name=name).exists():
            return JsonResponse({
                'success': False, 
                'error': 'A class with this name already exists.'
            }, status=400)
        
        if code != class_level.code and ClassLevel.objects.filter(code=code).exists():
            return JsonResponse({
                'success': False, 
                'error': 'A class with this code already exists.'
            }, status=400)
        
        class_level.name = name
        class_level.code = code
        class_level.description = description
        class_level.capacity = capacity
        class_level.display_order = display_order
        class_level.is_active = is_active
        class_level.form_teacher_id = form_teacher_id if form_teacher_id else None
        class_level.save()
        
        if subject_ids:
            if isinstance(subject_ids, str):
                import ast
                subject_ids = ast.literal_eval(subject_ids)
            subjects = Subject.objects.filter(id__in=subject_ids)
            class_level.subjects.set(subjects)
        
        response_data = {
            'success': True,
            'message': f'Class {name} updated successfully.',
        }
        
        if not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            messages.success(request, response_data['message'])
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'An error occurred while updating class: {str(e)}'
        }, status=400)



@login_required
@require_http_methods(["DELETE", "POST"])
def class_delete(request, class_id):
    """Delete a class"""
    try:
        class_level = get_object_or_404(ClassLevel, id=class_id)
        class_name = class_level.name
        
        # Check if class has students
        if class_level.students.exists():
            return JsonResponse({
                'success': False, 
                'error': f'Cannot delete {class_name} because it has students assigned. Please reassign students first.'
            }, status=400)
        
        class_level.delete()
        
        response_data = {
            'success': True,
            'message': f'Class {class_name} deleted successfully.',
        }
        
        if not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            messages.success(request, response_data['message'])
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'An error occurred while deleting class: {str(e)}'
        }, status=400)



@login_required
@require_http_methods(["GET"])
def get_subject_data(request, subject_id):
    """Get subject data for AJAX requests"""
    try:
        subject = get_object_or_404(Subject, id=subject_id)
        
        data = {
            'id': str(subject.id),
            'name': subject.name,
            'code': subject.code,
            'description': subject.description,
            'category': subject.category,
            'is_active': subject.is_active,
            'teacher': str(subject.teacher.id) if subject.teacher else None,
            'teacher_name': subject.teacher.get_full_name() if subject.teacher else None,
            'created_at': subject.created_at.isoformat(),
            'updated_at': subject.updated_at.isoformat(),
        }
        
        return JsonResponse({'success': True, 'data': data})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

@login_required
@require_http_methods(["GET"])
def get_class_data(request, class_id):
    """Get class data for AJAX requests"""
    try:
        class_level = get_object_or_404(ClassLevel, id=class_id)
        
        data = {
            'id': str(class_level.id),
            'name': class_level.name,
            'code': class_level.code,
            'description': class_level.description,
            'capacity': class_level.capacity,
            'display_order': class_level.display_order,
            'is_active': class_level.is_active,
            'form_teacher': str(class_level.form_teacher.id) if class_level.form_teacher else None,
            'form_teacher_name': class_level.form_teacher.get_full_name() if class_level.form_teacher else None,
            'current_students_count': class_level.current_students_count,
            'available_seats': class_level.available_seats,
            'subjects': list(class_level.subjects.values('id', 'name', 'code')),
            'created_at': class_level.created_at.isoformat(),
            'updated_at': class_level.updated_at.isoformat(),
        }
        
        return JsonResponse({'success': True, 'data': data})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@require_http_methods(["GET"])
def get_teachers_list(request):
    """Get list of teachers for dropdowns"""
    try:
        teachers = User.objects.filter(
            role='teacher',
            teacher_profile__is_active=True
        ).select_related('teacher_profile').values(
            'id', 'first_name', 'last_name', 'email', 'teacher_profile__employee_id'
        )
        
        teachers_list = list(teachers)
        return JsonResponse({'success': True, 'teachers': teachers_list})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@require_http_methods(["GET"])
def get_subjects_list(request):
    """Get list of subjects for dropdowns"""
    try:
        subjects = Subject.objects.filter(is_active=True).values('id', 'name', 'code', 'category')
        subjects_list = list(subjects)
        return JsonResponse({'success': True, 'subjects': subjects_list})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)



@login_required
@require_http_methods(["GET", "POST"])
def create_academic_year(request):
    """Create a new academic year"""
    if request.method == 'POST':
        try:
            data = request.POST
            
            name = data.get('name')
            start_date = data.get('start_date')
            end_date = data.get('end_date')
            is_current = data.get('is_current') == 'on'
            
            if not all([name, start_date, end_date]):
                return JsonResponse({
                    'success': False,
                    'error': 'All fields are required: name, start_date, end_date'
                }, status=400)
            

            start_date_obj = timezone.datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date_obj = timezone.datetime.strptime(end_date, '%Y-%m-%d').date()
            
            if start_date_obj >= end_date_obj:
                return JsonResponse({
                    'success': False,
                    'error': 'End date must be after start date'
                }, status=400)
            
            overlapping_years = AcademicYear.objects.filter(
                models.Q(start_date__lte=end_date_obj) & models.Q(end_date__gte=start_date_obj)
            )
            
            if overlapping_years.exists():
                return JsonResponse({
                    'success': False,
                    'error': 'Academic year overlaps with existing academic year'
                }, status=400)


            academic_year = AcademicYear(
                name=name,
                start_date=start_date_obj,
                end_date=end_date_obj,
                is_current=is_current
            )
            
            academic_year.save()
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': f'Academic year "{name}" created successfully',
                    'academic_year': {
                        'id': str(academic_year.id),
                        'name': academic_year.name,
                        'start_date': academic_year.start_date.isoformat(),
                        'end_date': academic_year.end_date.isoformat(),
                        'is_current': academic_year.is_current
                    }
                })
            else:
                messages.success(request, f'Academic year "{name}" created successfully')
                return redirect('academic_years_list')
                
        except ValueError as e:
            return JsonResponse({
                'success': False,
                'error': 'Invalid date format. Use YYYY-MM-DD'
            }, status=400)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error creating academic year: {str(e)}'
            }, status=400)
    
    return render(request, 'pages/academics/create_academic_year.html')


@login_required
def get_academic_years(request):
    """API endpoint to get all academic years for dropdown"""
    academic_years = AcademicYear.objects.all().order_by('-start_date')
    
    years_data = [
        {
            'id': str(year.id),
            'name': year.name,
            'is_current': year.is_current,
            'start_date': year.start_date.isoformat(),
            'end_date': year.end_date.isoformat()
        }
        for year in academic_years
    ]
    
    return JsonResponse({'academic_years': years_data})


@login_required
def academic_years_list(request):
    """List all academic years"""
    academic_years = AcademicYear.objects.all().order_by('-start_date')
    return render(request, 'pages/academics/academic_years_list.html', {
        'academic_years': academic_years
    })


@login_required
@require_http_methods(["POST"])
@transaction.atomic
def assign_subjects_to_class(request, class_id):
    """API endpoint for assigning subjects to a class"""
    try:
        class_level = get_object_or_404(ClassLevel, id=class_id)
        
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            data = request.POST
        
        subject_ids = data.getlist("subjects") if hasattr(data, 'getlist') else data.get("subjects", [])
        academic_year_id = data.get("academic_year")
        
        if isinstance(subject_ids, str):
            subject_ids = ast.literal_eval(subject_ids)

        try:
            academic_year = AcademicYear.objects.get(id=academic_year_id)
        except AcademicYear.DoesNotExist:
            return JsonResponse({
                'success': False, 
                'error': 'Selected academic year does not exist.'
            }, status=400)
        
        subjects = Subject.objects.filter(id__in=subject_ids)
        
        ClassSubject.objects.filter(
            class_level=class_level, 
            academic_year=academic_year
        ).delete()
        
        class_subjects = []
        for subject in subjects:
            class_subjects.append(
                ClassSubject(
                    class_level=class_level,
                    subject=subject,
                    academic_year=academic_year
                )
            )
        
        ClassSubject.objects.bulk_create(class_subjects)
        
        response_data = {
            'success': True,
            'message': f'Successfully updated subjects for {class_level.name}.',
            'subjects_count': subjects.count(),
        }
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'An error occurred while updating class subjects: {str(e)}'
        }, status=400)
    

@login_required
@require_http_methods(["POST"])
@transaction.atomic
def assign_teacher_to_class(request, class_id):
    """API endpoint for assigning teacher to a class"""
    try:
        class_level = get_object_or_404(ClassLevel, id=class_id)
        
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            data = request.POST
        
        teacher_id = data.get("form_teacher")
        
        class_level.form_teacher_id = teacher_id if teacher_id else None
        class_level.save()
        
        teacher_name = class_level.form_teacher.get_full_name() if class_level.form_teacher else "No teacher"
        
        response_data = {
            'success': True,
            'message': f'Successfully assigned {teacher_name} to {class_level.name}.',
        }
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'An error occurred while assigning teacher: {str(e)}'
        }, status=400)
    
@login_required
@require_http_methods(["POST"])
def assign_teacher_to_subject(request, subject_id):
    """API endpoint for assigning teacher to a subject"""
    try:
        subject = get_object_or_404(Subject, id=subject_id)
        
        teacher_id = request.POST.get("teacher_id")
        
        if teacher_id:
            teacher = get_object_or_404(User, id=teacher_id, role='teacher')
            subject.teacher = teacher
        else:
            subject.teacher = None
            
        subject.save()
        
        teacher_name = subject.teacher.get_full_name() if subject.teacher else "No teacher"
        
        response_data = {
            'success': True,
            'message': f'Successfully assigned {teacher_name} to {subject.name}.',
        }
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'An error occurred while assigning teacher: {str(e)}'
        }, status=400)


@login_required
def get_teachers_api(request):
    """API endpoint to get all teachers for dropdown"""
    teachers = User.objects.filter(role='teacher', is_active=True).order_by('first_name')
    
    teachers_data = [
        {
            'id': str(teacher.id),
            'name': teacher.get_full_name(),
            'email': teacher.email
        }
        for teacher in teachers
    ]
    
    return JsonResponse({
        'success': True,
        'teachers': teachers_data
    })



@login_required
def manage_subject_teachers(request, subject_id):
    """View to manage teachers for a subject across different classes"""
    subject = get_object_or_404(Subject, id=subject_id)
    
    class_assignments = ClassSubject.objects.filter(
        subject=subject
    ).select_related('class_level', 'teacher', 'academic_year')
    
    classes = ClassLevel.objects.filter(is_active=True)
    teachers = User.objects.filter(role='teacher', is_active=True)
    academic_years = AcademicYear.objects.all()
    
    context = {
        'subject': subject,
        'class_assignments': class_assignments,
        'classes': classes,
        'teachers': teachers,
        'academic_years': academic_years,
    }
    
    return render(request, 'pages/admin_dashboard/manage_subject_teachers.html', context)


# @login_required
# @require_http_methods(["POST"])
# def assign_teacher_to_class_subject(request):
#     """API endpoint for assigning teacher to a subject for a specific class"""
#     try:
#         subject_id = request.POST.get("subject_id")
#         class_level_id = request.POST.get("class_level_id")
#         teacher_id = request.POST.get("teacher_id")
#         academic_year_id = request.POST.get("academic_year_id")
        
#         subject = get_object_or_404(Subject, id=subject_id)
#         class_level = get_object_or_404(ClassLevel, id=class_level_id)
#         academic_year = get_object_or_404(AcademicYear, id=academic_year_id)
#         teacher = get_object_or_404(User, id=teacher_id, role='teacher') if teacher_id else None
        
#         class_subject, created = ClassSubject.objects.update_or_create(
#             class_level=class_level,
#             subject=subject,
#             academic_year=academic_year,
#             defaults={'teacher': teacher}
#         )
        
#         action = "assigned" if created else "updated"
#         teacher_name = teacher.get_full_name() if teacher else "No teacher"
        
#         response_data = {
#             'success': True,
#             'message': f'Successfully {action} {teacher_name} to {subject.name} for {class_level.name}.',
#         }
        
#         return JsonResponse(response_data)
        
#     except Exception as e:
#         return JsonResponse({
#             'success': False, 
#             'error': f'An error occurred while assigning teacher: {str(e)}'
#         }, status=400)
    


@login_required
@require_http_methods(["POST"])
def assign_teacher_to_class_subject(request):
    """API endpoint for assigning teacher to a subject for a specific class"""
    try:
        subject_id = request.POST.get("subject_id")
        class_level_id = request.POST.get("class_level_id")
        teacher_id = request.POST.get("teacher_id")
        academic_year_id = request.POST.get("academic_year_id")
        
        subject = get_object_or_404(Subject, id=subject_id)
        class_level = get_object_or_404(ClassLevel, id=class_level_id)
        academic_year = get_object_or_404(AcademicYear, id=academic_year_id)
        teacher = get_object_or_404(User, id=teacher_id, role='teacher') if teacher_id else None
        
        # Check if assignment already exists
        existing_assignment = ClassSubject.objects.filter(
            class_level=class_level,
            subject=subject,
            academic_year=academic_year
        ).first()
        
        if existing_assignment:
            existing_assignment.teacher = teacher
            existing_assignment.save()
            action = "updated"
        else:
            ClassSubject.objects.create(
                class_level=class_level,
                subject=subject,
                academic_year=academic_year,
                teacher=teacher
            )
            action = "added"
        
        teacher_name = teacher.get_full_name() if teacher else "No teacher"
        
        response_data = {
            'success': True,
            'message': f'Successfully {action} {teacher_name} to {subject.name} for {class_level.name}.',
        }
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'An error occurred while assigning teacher: {str(e)}'
        }, status=400)


@login_required
@require_http_methods(["POST"])
def update_class_assignment(request):
    """API endpoint for updating class assignment teacher"""
    try:
        assignment_id = request.POST.get("assignment_id")
        teacher_id = request.POST.get("teacher_id")
        
        assignment = get_object_or_404(ClassSubject, id=assignment_id)
        teacher = get_object_or_404(User, id=teacher_id, role='teacher') if teacher_id else None
        
        assignment.teacher = teacher
        assignment.save()
        
        teacher_name = teacher.get_full_name() if teacher else "No teacher"
        
        response_data = {
            'success': True,
            'message': f'Successfully updated teacher to {teacher_name} for {assignment.class_level.name} - {assignment.subject.name}.',
        }
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'An error occurred while updating assignment: {str(e)}'
        }, status=400)


@login_required
@require_http_methods(["DELETE"])
def delete_class_assignment(request, assignment_id):
    """API endpoint for deleting a class assignment"""
    try:
        assignment = get_object_or_404(ClassSubject, id=assignment_id)
        class_name = assignment.class_level.name
        subject_name = assignment.subject.name
        
        assignment.delete()
        
        response_data = {
            'success': True,
            'message': f'Successfully removed assignment for {class_name} - {subject_name}.',
        }
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'An error occurred while removing assignment: {str(e)}'
        }, status=400)  
    



@login_required
def results_dashboard(request):
    """Main results dashboard for teachers and admin"""
    
    # Get filter parameters
    academic_year_id = request.GET.get('academic_year')
    class_level_id = request.GET.get('class_level')
    term_id = request.GET.get('term')
    subject_id = request.GET.get('subject')
    
    # Base queryset
    if request.user.role == 'teacher':
        results = Result.objects.filter(
            Q(uploaded_by=request.user) | Q(subject__teacher=request.user)
        )
    else:
        results = Result.objects.all()
    
    # Apply filters
    if academic_year_id:
        results = results.filter(term__academic_year_id=academic_year_id)
    if class_level_id:
        results = results.filter(class_level_id=class_level_id)
    if term_id:
        results = results.filter(term_id=term_id)
    if subject_id:
        results = results.filter(subject_id=subject_id)
    
    # Get available filters
    academic_years = AcademicYear.objects.all().order_by('-start_date')
    class_levels = ClassLevel.objects.filter(is_active=True)
    terms = Term.objects.all().order_by('-academic_year', 'start_date')
    
    if request.user.role == 'teacher':
        subjects = Subject.objects.filter(
            Q(teacher=request.user) | Q(class_subjects__teacher=request.user)
        ).distinct()
    else:
        subjects = Subject.objects.all()
    
    # Statistics
    total_results = results.count()
    published_results = results.filter(is_published=True).count()
    average_score = results.aggregate(avg=Avg('score'))['avg'] or 0
    
    # Recent results
    recent_results = results.select_related(
        'student', 'subject', 'class_level', 'term'
    ).order_by('-date_uploaded')[:10]
    
    # Performance by subject
    subject_performance = results.values(
        'subject__name', 'subject__code'
    ).annotate(
        avg_score=Avg('score'),
        count=Count('id'),
        max_score=Max('score'),
        min_score=Min('score')
    ).order_by('-avg_score')
    
    context = {
        'total_results': total_results,
        'published_results': published_results,
        'average_score': round(average_score, 2),
        'recent_results': recent_results,
        'subject_performance': subject_performance,
        'academic_years': academic_years,
        'class_levels': class_levels,
        'terms': terms,
        'subjects': subjects,
        'current_filters': {
            'academic_year_id': academic_year_id,
            'class_level_id': class_level_id,
            'term_id': term_id,
            'subject_id': subject_id,
        }
    }
    
    return render(request, 'pages/admin_dashboard/results.html', {'context': context})


@login_required
def upload_results(request):
    """Upload results for a class and subject"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            class_level_id = data.get('class_level_id')
            subject_id = data.get('subject_id')
            term_id = data.get('term_id')
            results_data = data.get('results', [])
            
            class_level = get_object_or_404(ClassLevel, id=class_level_id)
            subject = get_object_or_404(Subject, id=subject_id)
            term = get_object_or_404(Term, id=term_id)
            
            # Check permissions
            if request.user.role == 'teacher':
                if not (subject.teacher == request.user or 
                       ClassSubject.objects.filter(
                           class_level=class_level,
                           subject=subject,
                           teacher=request.user
                       ).exists()):
                    return JsonResponse({
                        'success': False,
                        'error': 'You are not assigned to teach this subject for this class'
                    }, status=403)
            
            created_count = 0
            updated_count = 0
            
            with transaction.atomic():
                for result_data in results_data:
                    student_id = result_data.get('student_id')
                    score = result_data.get('score')
                    
                    if not student_id or score is None:
                        continue
                    
                    student = get_object_or_404(User, id=student_id, role='student')
                    
                    # Create or update result
                    result, created = Result.objects.update_or_create(
                        student=student,
                        subject=subject,
                        class_level=class_level,
                        term=term,
                        defaults={
                            'score': score,
                            'uploaded_by': request.user,
                            'is_published': False
                        }
                    )
                    
                    if created:
                        created_count += 1
                    else:
                        updated_count += 1
            
            return JsonResponse({
                'success': True,
                'message': f'Successfully processed {created_count + updated_count} results '
                          f'({created_count} created, {updated_count} updated)',
                'created_count': created_count,
                'updated_count': updated_count
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error uploading results: {str(e)}'
            }, status=400)
    
    # GET request - show upload form
    academic_years = AcademicYear.objects.all().order_by('-start_date')
    class_levels = ClassLevel.objects.filter(is_active=True)
    subjects = Subject.objects.filter(is_active=True)
    
    if request.user.role == 'teacher':
        subjects = subjects.filter(
            Q(teacher=request.user) | Q(class_subjects__teacher=request.user)
        ).distinct()
    
    context = {
        'academic_years': academic_years,
        'class_levels': class_levels,
        'subjects': subjects,
    }
    
    return render(request, 'academics/results/upload.html', {'context': context})


@login_required
def get_students_for_results(request):
    """Get students for a specific class to populate results form"""
    class_level_id = request.GET.get('class_level_id')
    academic_year_id = request.GET.get('academic_year_id')
    
    if not class_level_id:
        return JsonResponse({'success': False, 'error': 'Class level ID required'})
    
    class_level = get_object_or_404(ClassLevel, id=class_level_id)
    
    students = User.objects.filter(
        role='student',
        student_profile__current_class=class_level,
        student_profile__is_active=True
    ).select_related('student_profile').values(
        'id',
        'first_name',
        'last_name',
        'student_profile__student_id'
    ).order_by('first_name', 'last_name')
    
    return JsonResponse({
        'success': True,
        'students': list(students)
    })


@login_required
def get_existing_results(request):
    """Get existing results for a class, subject, and term"""
    class_level_id = request.GET.get('class_level_id')
    subject_id = request.GET.get('subject_id')
    term_id = request.GET.get('term_id')
    
    if not all([class_level_id, subject_id, term_id]):
        return JsonResponse({'success': False, 'error': 'All parameters required'})
    
    results = Result.objects.filter(
        class_level_id=class_level_id,
        subject_id=subject_id,
        term_id=term_id
    ).select_related('student').values(
        'student_id',
        'score',
        'grade',
        'is_published'
    )
    
    results_dict = {result['student_id']: result for result in results}
    
    return JsonResponse({
        'success': True,
        'results': results_dict
    })


@login_required
@require_http_methods(["POST"])
def publish_results(request, result_id=None):
    """Publish or unpublish results"""
    try:
        if result_id:
            # Publish single result
            result = get_object_or_404(Result, id=result_id)
            
            # Check permissions
            if request.user.role == 'teacher' and result.uploaded_by != request.user:
                return JsonResponse({
                    'success': False,
                    'error': 'You can only publish results you uploaded'
                }, status=403)
            
            result.is_published = not result.is_published
            result.save()
            
            action = 'published' if result.is_published else 'unpublished'
            
            return JsonResponse({
                'success': True,
                'message': f'Result {action} successfully',
                'is_published': result.is_published
            })
        else:
            # Bulk publish
            data = json.loads(request.body)
            result_ids = data.get('result_ids', [])
            publish = data.get('publish', True)
            
            results = Result.objects.filter(id__in=result_ids)
            
            # Check permissions for teachers
            if request.user.role == 'teacher':
                results = results.filter(uploaded_by=request.user)
            
            updated_count = results.update(is_published=publish)
            
            action = 'published' if publish else 'unpublished'
            
            return JsonResponse({
                'success': True,
                'message': f'{updated_count} results {action} successfully'
            })
            
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error publishing results: {str(e)}'
        }, status=400)


@login_required
def result_detail(request, result_id):
    """Detailed view of a single result"""
    result = get_object_or_404(
        Result.objects.select_related(
            'student', 'subject', 'class_level', 'term', 'uploaded_by'
        ),
        id=result_id
    )
    
    # Check permissions
    if request.user.role == 'student' and result.student != request.user:
        messages.error(request, 'You can only view your own results')
        return redirect('student_results')
    
    if request.user.role == 'teacher' and result.uploaded_by != request.user:
        messages.error(request, 'You can only view results you uploaded')
        return redirect('results_dashboard')
    
    context = {
        'result': result,
    }
    
    return render(request, 'pages/academics/results/detail.html', {'context': context})


@login_required
@require_http_methods(["DELETE"])
def delete_result(request, result_id):
    """Delete a result"""
    try:
        result = get_object_or_404(Result, id=result_id)
        
        if request.user.role == 'teacher' and result.uploaded_by != request.user:
            return JsonResponse({
                'success': False,
                'error': 'You can only delete results you uploaded'
            }, status=403)
        
        result.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Result deleted successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error deleting result: {str(e)}'
        }, status=400)


@login_required
def results_analysis(request):
    """Advanced results analysis"""
    # Get filter parameters
    academic_year_id = request.GET.get('academic_year')
    class_level_id = request.GET.get('class_level')
    term_id = request.GET.get('term')
    
    results = Result.objects.all()
    
    # Apply filters
    if academic_year_id:
        results = results.filter(term__academic_year_id=academic_year_id)
    if class_level_id:
        results = results.filter(class_level_id=class_level_id)
    if term_id:
        results = results.filter(term_id=term_id)
    
    # Comprehensive analysis
    # Grade distribution
    grade_distribution = results.values('grade').annotate(
        count=Count('id')
    ).order_by('grade')
    
    # Subject performance trends
    subject_trends = results.values(
        'subject__name',
        'term__name',
        'term__academic_year__name'
    ).annotate(
        avg_score=Avg('score'),
        student_count=Count('student', distinct=True)
    ).order_by('subject__name', 'term__academic_year__name', 'term__name')
    
    # Class performance comparison
    class_performance = results.values('class_level__name').annotate(
        avg_score=Avg('score'),
        total_students=Count('student', distinct=True),
        pass_rate=Count('id', filter=Q(score__gte=50)) * 100.0 / Count('id')
    ).order_by('-avg_score')
    
    # Top performers cut by 10
    top_performers = results.select_related('student', 'subject').order_by('-score')[:10]
    
    context = {
        'grade_distribution': list(grade_distribution),
        'subject_trends': list(subject_trends),
        'class_performance': list(class_performance),
        'top_performers': top_performers,
        'total_results_analyzed': results.count(),
    }
    
    return render(request, 'pages/academics/results/analysis.html', {'context': context})


@login_required
def get_terms_for_academic_year(request):
    """Get terms for a specific academic year"""
    academic_year_id = request.GET.get('academic_year')
    
    if not academic_year_id:
        return JsonResponse({'success': False, 'error': 'Academic year ID required'})
    
    terms = Term.objects.filter(academic_year_id=academic_year_id).values(
        'id', 'name', 'start_date', 'end_date'
    ).order_by('start_date')
    
    return JsonResponse({
        'success': True,
        'terms': list(terms)
    })






import os
import json
from datetime import datetime
from django.http import HttpResponse
from django.conf import settings
from django.template.loader import render_to_string
from django.db.models import Avg, Count, Q
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.cell.cell import MergedCell
import openpyxl
import io

@login_required
def export_analysis_report(request):
    """Export analysis report as PDF or Excel"""
    format_type = request.GET.get('format', 'pdf')
    academic_year_id = request.GET.get('academic_year')
    class_level_id = request.GET.get('class_level')
    term_id = request.GET.get('term')
    
    # Get filtered data
    analysis_data = get_analysis_data(academic_year_id, class_level_id, term_id)
    
    if format_type.lower() == 'excel':
        return export_excel_report(analysis_data, request)
    else:
        return export_pdf_report(analysis_data, request)

def get_analysis_data(academic_year_id=None, class_level_id=None, term_id=None):
    """Get comprehensive analysis data for export"""
    results = Result.objects.all()
    
    # Apply filters
    if academic_year_id:
        results = results.filter(term__academic_year_id=academic_year_id)
    if class_level_id:
        results = results.filter(class_level_id=class_level_id)
    if term_id:
        results = results.filter(term_id=term_id)
    
    # Comprehensive analysis data
    data = {
        'filters': {
            'academic_year': AcademicYear.objects.filter(id=academic_year_id).first() if academic_year_id else None,
            'class_level': ClassLevel.objects.filter(id=class_level_id).first() if class_level_id else None,
            'term': Term.objects.filter(id=term_id).first() if term_id else None,
        },
        'summary': {
            'total_results': results.count(),
            'total_students': results.values('student').distinct().count(),
            'average_score': results.aggregate(avg=Avg('score'))['avg'] or 0,
            'published_results': results.filter(is_published=True).count(),
        },
        'grade_distribution': list(results.values('grade').annotate(
            count=Count('id')
        ).order_by('grade')),
        'subject_performance': list(results.values(
            'subject__name', 'subject__code'
        ).annotate(
            avg_score=Avg('score'),
            total_students=Count('student', distinct=True),
            pass_count=Count('id', filter=Q(score__gte=50)),
            max_score=Max('score'),
            min_score=Min('score')
        ).order_by('-avg_score')),
        'class_performance': list(results.values('class_level__name').annotate(
            avg_score=Avg('score'),
            total_students=Count('student', distinct=True),
            pass_rate=Count('id', filter=Q(score__gte=50)) * 100.0 / Count('id')
        ).order_by('-avg_score')),
        'top_performers': list(results.select_related(
            'student', 'subject', 'class_level'
        ).order_by('-score')[:10].values(
            'student__first_name',
            'student__last_name',
            'student__student_profile__student_id',
            'subject__name',
            'class_level__name',
            'score',
            'grade'
        )),
        'generated_at': datetime.now()
    }
    
    return data

def export_pdf_report(analysis_data, request):
    """Generate PDF report"""
    try:
        # Create buffer for PDF
        buffer = io.BytesIO()
        
        # Create PDF document
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=72
        )
        
        # Get styles
        styles = getSampleStyleSheet()
        
        # Create custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1,  # Center
            textColor=colors.HexColor('#1e293b')
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=12,
            spaceBefore=20,
            textColor=colors.HexColor('#374151')
        )
        
        # Build story (content)
        story = []
        
        # Title
        title = Paragraph("Academic Results Analysis Report", title_style)
        story.append(title)
        
        # Report metadata
        metadata_style = ParagraphStyle(
            'Metadata',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#64748b')
        )
        
        filters_text = "Filters: "
        filters = []
        if analysis_data['filters']['academic_year']:
            filters.append(f"Academic Year: {analysis_data['filters']['academic_year'].name}")
        if analysis_data['filters']['class_level']:
            filters.append(f"Class: {analysis_data['filters']['class_level'].name}")
        if analysis_data['filters']['term']:
            filters.append(f"Term: {analysis_data['filters']['term'].name}")
        
        filters_text += " | ".join(filters) if filters else "All Data"
        filters_text += f" | Generated: {analysis_data['generated_at'].strftime('%Y-%m-%d %H:%M')}"
        
        metadata = Paragraph(filters_text, metadata_style)
        story.append(metadata)
        story.append(Spacer(1, 20))
        
        # Summary Section
        story.append(Paragraph("Executive Summary", heading_style))
        
        summary_data = [
            ['Metric', 'Value'],
            ['Total Results', f"{analysis_data['summary']['total_results']:,}"],
            ['Total Students', f"{analysis_data['summary']['total_students']:,}"],
            ['Average Score', f"{analysis_data['summary']['average_score']:.1f}%"],
            ['Published Results', f"{analysis_data['summary']['published_results']:,}"],
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafc')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0'))
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Grade Distribution
        story.append(Paragraph("Grade Distribution", heading_style))
        
        grade_data = [['Grade', 'Count', 'Percentage']]
        total_results = analysis_data['summary']['total_results']
        
        for grade in analysis_data['grade_distribution']:
            percentage = (grade['count'] / total_results * 100) if total_results > 0 else 0
            grade_data.append([
                grade['grade'],
                str(grade['count']),
                f"{percentage:.1f}%"
            ])
        
        grade_table = Table(grade_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch])
        grade_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d1d5db'))
        ]))
        story.append(grade_table)
        story.append(Spacer(1, 20))
        
        # Subject Performance
        story.append(Paragraph("Subject Performance", heading_style))
        
        subject_data = [['Subject', 'Avg Score', 'Students', 'Pass Rate', 'High Score']]
        
        for subject in analysis_data['subject_performance']:
            pass_rate = (subject['pass_count'] / subject['total_students'] * 100) if subject['total_students'] > 0 else 0
            subject_data.append([
                subject['subject__name'],
                f"{subject['avg_score']:.1f}%",
                str(subject['total_students']),
                f"{pass_rate:.1f}%",
                f"{subject['max_score']:.1f}%"
            ])
        
        subject_table = Table(subject_data, colWidths=[2*inch, 1*inch, 1*inch, 1*inch, 1*inch])
        subject_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f59e0b')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d1d5db'))
        ]))
        story.append(subject_table)
        story.append(Spacer(1, 20))
        
        # Class Performance
        story.append(Paragraph("Class Performance", heading_style))
        
        class_data = [['Class', 'Avg Score', 'Students', 'Pass Rate']]
        
        for class_perf in analysis_data['class_performance']:
            class_data.append([
                class_perf['class_level__name'],
                f"{class_perf['avg_score']:.1f}%",
                str(class_perf['total_students']),
                f"{class_perf['pass_rate']:.1f}%"
            ])
        
        class_table = Table(class_data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1.5*inch])
        class_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8b5cf6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d1d5db'))
        ]))
        story.append(class_table)
        story.append(Spacer(1, 20))
        
        # Top Performers
        story.append(Paragraph("Top 10 Performers", heading_style))
        
        if analysis_data['top_performers']:
            top_data = [['Student', 'Student ID', 'Subject', 'Class', 'Score', 'Grade']]
            
            for performer in analysis_data['top_performers']:
                top_data.append([
                    f"{performer['student__first_name']} {performer['student__last_name']}",
                    performer['student__student_profile__student_id'],
                    performer['subject__name'],
                    performer['class_level__name'],
                    f"{performer['score']:.1f}%",
                    performer['grade']
                ])
            
            top_table = Table(top_data, colWidths=[1.5*inch, 1*inch, 1.2*inch, 1*inch, 0.8*inch, 0.8*inch])
            top_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ef4444')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d1d5db'))
            ]))
            story.append(top_table)
        else:
            story.append(Paragraph("No top performers data available.", styles['Normal']))
        
        # Build PDF
        doc.build(story)
        
        # Get PDF value from buffer
        pdf = buffer.getvalue()
        buffer.close()
        
        # Create HTTP response
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="results_analysis_report.pdf"'
        response.write(pdf)
        
        return response
        
    except Exception as e:
        return HttpResponse(f"Error generating PDF: {str(e)}", status=500)


def export_excel_report(analysis_data, request):
    """Generate Excel report"""
    try:
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Summary Sheet
        summary_sheet = wb.create_sheet("Executive Summary")
        
        # Header
        summary_sheet.merge_cells('A1:D1')
        summary_sheet['A1'] = "Academic Results Analysis Report"
        summary_sheet['A1'].font = Font(size=16, bold=True, color="1e293b")
        summary_sheet['A1'].alignment = Alignment(horizontal='center')
        
        # Filters info
        filters_text = "Filters: "
        filters = []
        if analysis_data['filters']['academic_year']:
            filters.append(f"Academic Year: {analysis_data['filters']['academic_year'].name}")
        if analysis_data['filters']['class_level']:
            filters.append(f"Class: {analysis_data['filters']['class_level'].name}")
        if analysis_data['filters']['term']:
            filters.append(f"Term: {analysis_data['filters']['term'].name}")
        
        filters_text += " | ".join(filters) if filters else "All Data"
        summary_sheet['A3'] = filters_text
        summary_sheet['A4'] = f"Generated: {analysis_data['generated_at'].strftime('%Y-%m-%d %H:%M')}"
        
        # Summary Table
        summary_sheet['A6'] = "Metric"
        summary_sheet['B6'] = "Value"
        
        summary_data = [
            ['Total Results', analysis_data['summary']['total_results']],
            ['Total Students', analysis_data['summary']['total_students']],
            ['Average Score', analysis_data['summary']['average_score']],
            ['Published Results', analysis_data['summary']['published_results']],
        ]
        
        for i, (metric, value) in enumerate(summary_data, start=7):
            summary_sheet[f'A{i}'] = metric
            summary_sheet[f'B{i}'] = value
        
        # Style summary table
        for row in summary_sheet['A6:B10']:
            for cell in row:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="f8fafc", end_color="f8fafc", fill_type="solid")
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )
        
        # Grade Distribution Sheet
        grade_sheet = wb.create_sheet("Grade Distribution")
        
        grade_sheet['A1'] = "Grade"
        grade_sheet['B1'] = "Count"
        grade_sheet['C1'] = "Percentage"
        
        for i, grade in enumerate(analysis_data['grade_distribution'], start=2):
            percentage = (grade['count'] / analysis_data['summary']['total_results'] * 100) if analysis_data['summary']['total_results'] > 0 else 0
            grade_sheet[f'A{i}'] = grade['grade']
            grade_sheet[f'B{i}'] = grade['count']
            grade_sheet[f'C{i}'] = percentage / 100  # Excel percentage format
        
        # Style grade sheet
        for row in grade_sheet['A1:C1']:
            for cell in row:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
        
        # Format percentage column
        for cell in grade_sheet['C']:
            if cell.row > 1:
                cell.number_format = '0.0%'
        
        # Subject Performance Sheet
        subject_sheet = wb.create_sheet("Subject Performance")
        
        subject_headers = ['Subject', 'Code', 'Avg Score', 'Students', 'Pass Count', 'Pass Rate', 'High Score', 'Low Score']
        for col, header in enumerate(subject_headers, start=1):
            subject_sheet.cell(row=1, column=col, value=header)
        
        for i, subject in enumerate(analysis_data['subject_performance'], start=2):
            pass_rate = (subject['pass_count'] / subject['total_students']) if subject['total_students'] > 0 else 0
            subject_sheet.cell(row=i, column=1, value=subject['subject__name'])
            subject_sheet.cell(row=i, column=2, value=subject['subject__code'])
            subject_sheet.cell(row=i, column=3, value=subject['avg_score'])
            subject_sheet.cell(row=i, column=4, value=subject['total_students'])
            subject_sheet.cell(row=i, column=5, value=subject['pass_count'])
            subject_sheet.cell(row=i, column=6, value=pass_rate)
            subject_sheet.cell(row=i, column=7, value=subject['max_score'])
            subject_sheet.cell(row=i, column=8, value=subject.get('min_score', 0))
        
        # Style subject sheet header
        for col in range(1, len(subject_headers) + 1):
            cell = subject_sheet.cell(row=1, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="f59e0b", end_color="f59e0b", fill_type="solid")
        
        # Format percentage and score columns
        for row in range(2, len(analysis_data['subject_performance']) + 2):
            subject_sheet.cell(row=row, column=6).number_format = '0.0%'
            subject_sheet.cell(row=row, column=3).number_format = '0.0'
            subject_sheet.cell(row=row, column=7).number_format = '0.0'
            subject_sheet.cell(row=row, column=8).number_format = '0.0'
        
        # Class Performance Sheet
        class_sheet = wb.create_sheet("Class Performance")
        
        class_headers = ['Class', 'Avg Score', 'Students', 'Pass Rate']
        for col, header in enumerate(class_headers, start=1):
            class_sheet.cell(row=1, column=col, value=header)
        
        for i, class_perf in enumerate(analysis_data['class_performance'], start=2):
            class_sheet.cell(row=i, column=1, value=class_perf['class_level__name'])
            class_sheet.cell(row=i, column=2, value=class_perf['avg_score'])
            class_sheet.cell(row=i, column=3, value=class_perf['total_students'])
            class_sheet.cell(row=i, column=4, value=class_perf['pass_rate'] / 100)
        
        # Style class sheet
        for col in range(1, len(class_headers) + 1):
            cell = class_sheet.cell(row=1, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="8b5cf6", end_color="8b5cf6", fill_type="solid")
        
        # Format columns
        for row in range(2, len(analysis_data['class_performance']) + 2):
            class_sheet.cell(row=row, column=2).number_format = '0.0'
            class_sheet.cell(row=row, column=4).number_format = '0.0%'
        
        # Top Performers Sheet
        if analysis_data['top_performers']:
            top_sheet = wb.create_sheet("Top Performers")
            
            top_headers = ['Rank', 'Student Name', 'Student ID', 'Subject', 'Class', 'Score', 'Grade']
            for col, header in enumerate(top_headers, start=1):
                top_sheet.cell(row=1, column=col, value=header)
            
            for i, performer in enumerate(analysis_data['top_performers'], start=2):
                top_sheet.cell(row=i, column=1, value=i-1)
                top_sheet.cell(row=i, column=2, value=f"{performer['student__first_name']} {performer['student__last_name']}")
                top_sheet.cell(row=i, column=3, value=performer['student__student_profile__student_id'])
                top_sheet.cell(row=i, column=4, value=performer['subject__name'])
                top_sheet.cell(row=i, column=5, value=performer['class_level__name'])
                top_sheet.cell(row=i, column=6, value=performer['score'])
                top_sheet.cell(row=i, column=7, value=performer['grade'])
            
            # Style top performers sheet
            for col in range(1, len(top_headers) + 1):
                cell = top_sheet.cell(row=1, column=col)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="ef4444", end_color="ef4444", fill_type="solid")
            
            # Format score column
            for row in range(2, len(analysis_data['top_performers']) + 2):
                top_sheet.cell(row=row, column=6).number_format = '0.0'
        
        # Auto-adjust column widths
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for col in ws.columns:
                max_length = 0
                col_letter = None
                
                for cell in col:
                    # Skip merged cells
                    if isinstance(cell, MergedCell):
                        continue
                    
                    if col_letter is None:
                        col_letter = cell.column_letter
                    
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                if col_letter:
                    ws.column_dimensions[col_letter].width = (max_length + 2) * 1.2
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Create HTTP response
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="results_analysis_report.xlsx"'
        
        return response
        
    except Exception as e:
        return HttpResponse(f"Error generating Excel report: {str(e)}", status=500)